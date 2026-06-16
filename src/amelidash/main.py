import pandas as pd
import os
from openpyxl import Workbook
from config import OUTPUT_DASHBOARD, SHEET_CLEANED_DEPENSES, SHEET_CLEANED_EFFECTIFS
from data import get_cleaned_effectifs, get_cleaned_depenses
from sheets.filtersdep import OngletFiltres
from sheets.filterseff import OngletFiltreseff
from sheets.graphiques import OngletGraphiques
from sheets.depenses import OngletDepenses
from sheets.departement import OngletDepartement
from sheets.postes import OngletPostes
from sheets.Analyse import OngletAnalysesUnifiees

"""Point d’entrée du script générant le dashboard Excel complet.

Ce module :

- charge et nettoie les données d’effectifs et de dépenses ;
- agrège les effectifs (suppression du niveau patho_niv2) pour optimiser l’écriture Excel ;
- écrit les données nettoyées dans les onglets cleanedData ;
- génère successivement tous les onglets du dashboard :
    • filtres dépenses  
    • filtres effectifs  
    • analyses dépenses  
    • graphiques effectifs  
    • postes de dépense  
    • département  
    • analyses unifiées  
- sauvegarde enfin le fichier Excel final.

Fournit également deux utilitaires :
- `write_df_to_sheet` : écrit un DataFrame en remplaçant tous les manquants par des cellules vides ;
- `reduce_effectifs` : agrège les effectifs au niveau patho_niv1 pour alléger les données.

Le script peut être exécuté directement pour produire le dashboard."""


def write_df_to_sheet(ws, df):
    """Écrit le DataFrame en convertissant tous les manquants (<NA>, NaN, None) en cellule vide."""
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        clean = []
        for v in row:
            if v is None or pd.isna(v):
                clean.append(None)
            elif isinstance(v, str) and v.strip().lower() == "nan":
                clean.append(None)
            else:
                clean.append(v)
        ws.append(clean)


def reduce_effectifs(df):
    """Agrège les effectifs en supprimant patho_niv2 (qui multiplie les lignes).
    Conserve le niveau patho_niv1 utilisé par les graphiques."""
    group_cols = [
        c
        for c in [
            "annee",
            "Sexe",
            "patho_niv1",
            "Classe d'age",
            "Région",
            "Département",
        ]
        if c in df.columns
    ]
    num_cols = [
        c
        for c in ["Effectif", "Population de référence", "Prévalence"]
        if c in df.columns
    ]
    return df.groupby(group_cols, observed=True)[num_cols].sum().reset_index()


def main():
    print("Dashboard Ameli")
    print("=" * 70)

    try:
        print("\nChargement des donnees...")
        df_effectifs = get_cleaned_effectifs()
        df_depenses = get_cleaned_depenses()

        # Version réduite pour l'écriture dans Excel (patho_niv2 agrégé)
        df_eff_reduit = reduce_effectifs(df_effectifs)

        print(f"   Effectifs: {len(df_effectifs)} lignes (total)")
        print(f"   Effectifs reduit: {len(df_eff_reduit)} lignes")
        print(f"   Depenses: {len(df_depenses)} lignes")

        print(f"\nCreation de {OUTPUT_DASHBOARD.name}...")
        wb = Workbook()
        wb.remove(wb.active)

        print("   - cleanedData_Depenses...")
        ws_dep = wb.create_sheet(SHEET_CLEANED_DEPENSES, 0)
        write_df_to_sheet(ws_dep, df_depenses)

        print("   - cleanedData_Effectifs...")
        ws_eff = wb.create_sheet(SHEET_CLEANED_EFFECTIFS, 1)
        write_df_to_sheet(ws_eff, df_eff_reduit)

        print("   - Filtres depenses...")
        OngletFiltres(wb, df_depenses).create_depenses()

        print("   - Filtres effectifs...")
        OngletFiltreseff(wb, df_eff_reduit).create_effectifs()

        print("   - Depenses...")
        OngletDepenses(wb, df_depenses).create()

        print("   - Graphiques effectifs...")
        OngletGraphiques(wb, df_eff_reduit).create()

        print("   - Postes...")
        OngletPostes(wb, df_depenses).create()

        print("   - Departement...")
        OngletDepartement(wb, df_eff_reduit).create()

        print("   - Analyses unifiees...")
        OngletAnalysesUnifiees(wb, df_depenses, df_eff_reduit).create()
        os.makedirs(os.path.dirname(OUTPUT_DASHBOARD), exist_ok=True)
        wb.save(OUTPUT_DASHBOARD)

        print(f"\nSucces - Fichier cree : {OUTPUT_DASHBOARD}")
        return 0

    except Exception as e:
        print(f"\nErreur: {e}")
        import traceback

        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())
