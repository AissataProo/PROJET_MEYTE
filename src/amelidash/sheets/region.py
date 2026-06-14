"""Onglet Region : tableau effectifs par région avec graphique."""

from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from config import COULEURS, SHEET_CLEANED_EFFECTIFS


class OngletRegion:

    def __init__(self, wb, df):
        self.wb = wb
        self.df = df
        self.sheet_cleaned = SHEET_CLEANED_EFFECTIFS

    def create(self):
        ws = self.wb.create_sheet("Region")
        ws.sheet_view.showGridLines = False

        ws["A1"] = "Effectifs par Région"
        ws["A1"].font = Font(bold=True, size=13, color=COULEURS["blanc"])
        ws["A1"].fill = PatternFill(
            start_color=COULEURS["principal"], fill_type="solid"
        )

        regions = sorted(self.df["Région"].dropna().unique())[:100]

        for idx, region in enumerate(regions, start=2):
            ws.cell(idx, 1).value = region
            ws.cell(idx, 2).value = (
                f"=SUMIF('{self.sheet_cleaned}'!A:A,A{idx},"
                f"'{self.sheet_cleaned}'!E:E)"
            )
            ws.cell(idx, 2).number_format = "#,##0"

        chart = BarChart()
        chart.type = "col"
        chart.title = "Effectifs par Région"
        chart.height = 12
        chart.width = 20

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(regions) + 1)
        cat_ref = Reference(ws, min_col=1, min_row=2, max_row=len(regions) + 1)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.series[0].graphicalProperties.solidFill = COULEURS["principal"][2:]
        ws.add_chart(chart, "D2")

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 16
