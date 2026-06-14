from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, Reference


class OngletDepenses:
    """Génère l’onglet Excel dédié aux dépenses : agrégation par pathologie,
    analyse comparative entre années et création d’un graphique en camembert
    pour visualiser la répartition des montants."""

    def __init__(self, wb, df_depenses):
        self.wb = wb
        self.df_depenses = df_depenses
        self.col_annee = self._find_col("annee")
        self.col_patho = self._find_col("patho")
        self.col_poste = self._find_col("poste")
        self.col_montant = self._find_col("montant")
        self.col_effectif = self._find_col("effectif")

    def _find_col(self, keyword):
        for col in self.df_depenses.columns:
            if keyword.lower() in col.lower():
                return col
        return None

    def create(self):
        if not self.col_annee or not self.col_patho or not self.col_montant:
            print(
                f"Colonnes manquantes. Disponibles: {self.df_depenses.columns.tolist()}"
            )
            return

        sheet_raw = None
        for name in self.wb.sheetnames:
            if "cleaned" in name.lower() or "depense" in name.lower():
                sheet_raw = name
                break

        if not sheet_raw:
            print(f"Pas de feuille trouvée. Disponibles: {self.wb.sheetnames}")
            return

        ws_raw = self.wb[sheet_raw]

        headers = {}
        for idx, cell in enumerate(ws_raw[1], start=1):
            if cell.value is not None:
                headers[str(cell.value).strip()] = idx

        col_annee = headers.get(self.col_annee, 1) if self.col_annee else 1
        col_patho = headers.get(self.col_patho, 2) if self.col_patho else 2
        col_poste = headers.get(self.col_poste, 5) if self.col_poste else 5
        col_montant = headers.get(self.col_montant, 7) if self.col_montant else 7
        col_effectif = headers.get(self.col_effectif, 8) if self.col_effectif else 8

        col_annee_letter = get_column_letter(col_annee)
        col_patho_letter = get_column_letter(col_patho)
        col_poste_letter = get_column_letter(col_poste)
        col_montant_letter = get_column_letter(col_montant)
        col_effectif_letter = get_column_letter(col_effectif)

        annees_list = sorted(
            self.df_depenses[self.col_annee].dropna().unique(), reverse=True
        )
        postes_list = sorted(
            [
                str(v).strip()
                for v in self.df_depenses[self.col_poste].dropna().unique()
                if "total" not in str(v).lower()
            ]
        )

        if not annees_list:
            annees_list = [2023]
        if not postes_list:
            postes_list = ["Poste 1"]

        poste_defaut = postes_list[0]
        annee_defaut = int(annees_list[0])
        annee_n = int(annees_list[0])
        annee_n1 = int(annees_list[1]) if len(annees_list) >= 2 else 2022

        if "Dépenses" in self.wb.sheetnames:
            del self.wb["Dépenses"]

        ws = self.wb.create_sheet("Dépenses", 1)
        ws.sheet_view.showGridLines = False

        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        couleur_alt_1 = PatternFill(start_color="FFF5F5F5", fill_type="solid")
        couleur_alt_2 = PatternFill(start_color="FFFFFFFF", fill_type="solid")

        COLOR_VERT = "FF006B4F"
        COLOR_BLEU = "FF4472C4"
        COLOR_VERT2 = "FF70AD47"

        def style_data_cell(cell, idx=0, num_format=None, align="left"):
            cell.fill = couleur_alt_1 if idx % 2 == 0 else couleur_alt_2
            cell.font = Font(color="000000")
            cell.border = border
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if num_format:
                cell.number_format = num_format

        def make_header_cell(cell, label, color=COLOR_BLEU):
            cell.value = label
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin", color=color),
                right=Side(style="thin", color=color),
                top=Side(style="thin", color=color),
                bottom=Side(style="thin", color=color),
            )

        ws.merge_cells("A1:H1")
        ws["A1"] = "Dépenses par pathologie"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color=COLOR_VERT, fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws["A3"] = "Poste"
        ws["A3"].font = Font(bold=True, color="FFFFFF", size=11)
        ws["A3"].fill = PatternFill(start_color=COLOR_VERT, fill_type="solid")

        ws["B3"] = poste_defaut
        ws["B3"].font = Font(size=11, bold=True)
        ws["B3"].fill = PatternFill(start_color="FFFFFFFF", fill_type="solid")

        postes_str = ",".join(postes_list[:30])
        dv_poste = DataValidation(
            type="list", formula1=f'"{postes_str}"', allow_blank=False
        )
        ws.add_data_validation(dv_poste)
        dv_poste.add("B3")

        ws["D3"] = "Année"
        ws["D3"].font = Font(bold=True, color="FFFFFF", size=11)
        ws["D3"].fill = PatternFill(start_color=COLOR_VERT, fill_type="solid")

        ws["E3"] = annee_defaut
        ws["E3"].font = Font(size=11, bold=True)
        ws["E3"].fill = PatternFill(start_color="FFFFFFFF", fill_type="solid")

        annees_str = ",".join(str(int(a)) for a in annees_list)
        dv_annee = DataValidation(
            type="list", formula1=f'"{annees_str}"', allow_blank=False
        )
        ws.add_data_validation(dv_annee)
        dv_annee.add("E3")

        ws.row_dimensions[3].height = 22

        patho_rows = []
        seen = set()

        for row in ws_raw.iter_rows(min_row=2, values_only=True):
            try:
                val_annee = row[col_annee - 1]
                val_poste = row[col_poste - 1]
                val_patho = row[col_patho - 1]

                if val_annee and val_poste and val_patho:
                    # Vérification des valeurs
                    if (
                        int(val_annee) == int(annee_defaut)
                        and str(val_poste).strip() == str(poste_defaut).strip()
                    ):
                        label = str(val_patho).strip()

                        # Logique d'exclusion
                        exclude = (
                            "total" in label.lower()
                            or "soins courants" in label.lower()
                        )

                        if not exclude and label not in seen:
                            seen.add(label)
                            patho_rows.append(label)
            except Exception:
                pass

        ws.merge_cells("A5:D5")
        ws["A5"] = "Détails des pathologies"
        ws["A5"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A5"].fill = PatternFill(start_color=COLOR_VERT, fill_type="solid")
        ws.row_dimensions[5].height = 20

        TABLE1_HDR = 6
        TABLE1_START = TABLE1_HDR + 1

        make_header_cell(ws.cell(TABLE1_HDR, 1), "Pathologie", COLOR_BLEU)
        make_header_cell(ws.cell(TABLE1_HDR, 2), "Effectifs", COLOR_BLEU)
        make_header_cell(ws.cell(TABLE1_HDR, 3), "Montant (€)", COLOR_BLEU)
        make_header_cell(ws.cell(TABLE1_HDR, 4), "Coût/patient", COLOR_BLEU)
        ws.row_dimensions[TABLE1_HDR].height = 22

        for idx, label in enumerate(patho_rows):
            r = TABLE1_START + idx

            ws.cell(r, 1).value = label
            style_data_cell(ws.cell(r, 1), idx, align="left")

            ws.cell(r, 2).value = (
                f"=IFERROR(SUMIFS('{sheet_raw}'!${col_effectif_letter}:${col_effectif_letter},"
                f"'{sheet_raw}'!${col_patho_letter}:${col_patho_letter},A{r},"
                f"'{sheet_raw}'!${col_poste_letter}:${col_poste_letter},B$3,"
                f"'{sheet_raw}'!${col_annee_letter}:${col_annee_letter},E$3),0)"
            )
            style_data_cell(ws.cell(r, 2), idx, num_format="#,##0", align="right")

            ws.cell(r, 3).value = (
                f"=IFERROR(SUMIFS('{sheet_raw}'!${col_montant_letter}:${col_montant_letter},"
                f"'{sheet_raw}'!${col_patho_letter}:${col_patho_letter},A{r},"
                f"'{sheet_raw}'!${col_poste_letter}:${col_poste_letter},B$3,"
                f"'{sheet_raw}'!${col_annee_letter}:${col_annee_letter},E$3),0)"
            )
            style_data_cell(ws.cell(r, 3), idx, num_format="#,##0 €", align="right")

            ws.cell(r, 4).value = f"=IFERROR(C{r}/B{r},0)"
            style_data_cell(ws.cell(r, 4), idx, num_format="#,##0.00 €", align="right")

            ws.row_dimensions[r].height = 18

        TABLE1_END = TABLE1_START + len(patho_rows) - 1 if patho_rows else TABLE1_START

        # --- AJOUT DU GRAPHIQUE PIE (agrandi, sans légende) ---
        if patho_rows:
            pie = PieChart()
            pie.title = "Coût/Patient par Pathologie"
            pie.height = 16
            pie.width = 24
            pie.legend = None  # enlève la légende (mets une vraie légende en commentant cette ligne)
            # Données : Colonne 4 (Coût/patient)
            data_pie = Reference(ws, min_col=4, min_row=TABLE1_HDR, max_row=TABLE1_END)
            # Catégories : Colonne 1 (Pathologie)
            cats_pie = Reference(
                ws, min_col=1, min_row=TABLE1_START, max_row=TABLE1_END
            )

            pie.add_data(data_pie, titles_from_data=True)
            pie.set_categories(cats_pie)
            ws.add_chart(pie, "F6")

        S2 = TABLE1_END + 3

        ws.merge_cells(f"A{S2}:E{S2}")
        ws[f"A{S2}"] = "Comparaison annuelle"
        ws[f"A{S2}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{S2}"].fill = PatternFill(start_color=COLOR_VERT, fill_type="solid")
        ws[f"A{S2}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[S2].height = 25

        COMP_HDR = S2 + 1
        COMP_START = COMP_HDR + 1

        make_header_cell(ws.cell(COMP_HDR, 1), "Pathologie", COLOR_VERT2)
        make_header_cell(ws.cell(COMP_HDR, 2), f"Dép. {annee_n1}", COLOR_VERT2)
        make_header_cell(ws.cell(COMP_HDR, 3), f"Dép. {annee_n}", COLOR_VERT2)
        make_header_cell(ws.cell(COMP_HDR, 4), "Variation", COLOR_VERT2)
        make_header_cell(ws.cell(COMP_HDR, 5), "Évol. %", COLOR_VERT2)
        ws.row_dimensions[COMP_HDR].height = 22

        for idx, label in enumerate(patho_rows):
            r = COMP_START + idx

            ws.cell(r, 1).value = label
            style_data_cell(ws.cell(r, 1), idx, align="left")

            ws.cell(r, 2).value = (
                f"=IFERROR(SUMIFS('{sheet_raw}'!${col_montant_letter}:${col_montant_letter},"
                f"'{sheet_raw}'!${col_patho_letter}:${col_patho_letter},A{r},"
                f"'{sheet_raw}'!${col_poste_letter}:${col_poste_letter},B$3,"
                f"'{sheet_raw}'!${col_annee_letter}:${col_annee_letter},{annee_n1}),0)"
            )
            style_data_cell(ws.cell(r, 2), idx, num_format="#,##0 €", align="right")

            ws.cell(r, 3).value = (
                f"=IFERROR(SUMIFS('{sheet_raw}'!${col_montant_letter}:${col_montant_letter},"
                f"'{sheet_raw}'!${col_patho_letter}:${col_patho_letter},A{r},"
                f"'{sheet_raw}'!${col_poste_letter}:${col_poste_letter},B$3,"
                f"'{sheet_raw}'!${col_annee_letter}:${col_annee_letter},{annee_n}),0)"
            )
            style_data_cell(ws.cell(r, 3), idx, num_format="#,##0 €", align="right")

            ws.cell(r, 4).value = f"=C{r}-B{r}"
            style_data_cell(ws.cell(r, 4), idx, num_format="#,##0 €", align="right")

            ws.cell(r, 5).value = f"=IFERROR((C{r}-B{r})/B{r},0)"
            style_data_cell(ws.cell(r, 5), idx, num_format="0.00%", align="right")

            ws.row_dimensions[r].height = 18

        ws.column_dimensions["A"].width = 65
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 16


def create_onglet_depenses(wb, df_depenses):
    onglet = OngletDepenses(wb, df_depenses)
    onglet.create()
