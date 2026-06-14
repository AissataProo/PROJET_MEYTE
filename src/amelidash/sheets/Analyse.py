"""Onglet Analyses Unifiées - Top 10 Effectifs + Top 10 Montants, filtre Année dynamique."""

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import COULEURS, SHEET_CLEANED_DEPENSES


class OngletAnalysesUnifiees:
    def __init__(self, wb, df_depenses, df_effectifs):
        self.wb = wb
        self.df_depenses = df_depenses.copy()
        self.df_effectifs = df_effectifs.copy()
        self.couleur_principale = COULEURS["principal"]
        self.couleur_accent = COULEURS["accent"]

    def create(self):
        for nom in ["Analyses", "Data_Analyses"]:
            if nom in self.wb.sheetnames:
                self.wb.remove(self.wb[nom])

        ws = self.wb.create_sheet("Analyses")
        ws.sheet_view.showGridLines = False
        ws_data = self.wb.create_sheet("Data_Analyses")
        ws_data.sheet_state = "hidden"

        dep = self.df_depenses
        eff = self.df_effectifs
        dep["montant"] = pd.to_numeric(dep["montant"], errors="coerce").fillna(0)
        eff["Effectif"] = pd.to_numeric(eff["Effectif"], errors="coerce").fillna(0)

        annees = sorted(dep["annee"].dropna().unique(), reverse=True)
        annee_defaut = int(annees[0]) if annees else 2023

        # --- Titre ---
        ws.merge_cells("A1:T1")
        ws["A1"] = "Analyses unifiées"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=self.couleur_principale)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # --- Filtre Année (liste courte -> en dur, OK) ---
        ws["A3"] = "Année"
        ws["A3"].font = Font(bold=True, color="FFFFFF")
        ws["A3"].fill = PatternFill("solid", fgColor=self.couleur_accent)
        ws["A3"].alignment = Alignment(horizontal="center")
        ws["B3"] = annee_defaut
        ws["B3"].font = Font(bold=True)
        ws["B3"].alignment = Alignment(horizontal="center")
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(str(int(a)) for a in annees)}"',
            allow_blank=False,
        )
        ws.add_data_validation(dv)
        dv.add("B3")
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12

        # ===== Source EFFECTIF agrégée (patho x annee) =====
        sexes = set(eff["Sexe"].astype(str).unique())
        base_eff = eff[eff["Sexe"].astype(str) == "ENSEMBLE"] if "ENSEMBLE" in sexes else eff
        base_eff = base_eff[
            ~base_eff["Classe d'age"].astype(str).str.lower().str.contains("tous")
            & ~base_eff["patho_niv1"].astype(str).str.lower().str.contains("total")
        ]
        eff_agg = (
            base_eff.groupby(["patho_niv1", "annee"], observed=True)["Effectif"]
            .sum()
            .reset_index()
        )
        # écrit en colonnes H/I/J de Data_Analyses
        ws_data.cell(1, 8, "patho")
        ws_data.cell(1, 9, "annee")
        ws_data.cell(1, 10, "effectif")
        for i, r in enumerate(eff_agg.itertuples(index=False), start=2):
            ws_data.cell(i, 8, str(r[0]))
            ws_data.cell(i, 9, int(r[1]))
            ws_data.cell(i, 10, float(r[2]))

        # ===== Top 10 pathologies (effectif, année par défaut) =====
        top_eff = (
            eff_agg[eff_agg["annee"] == annee_defaut]
            .groupby("patho_niv1")["Effectif"]
            .sum()
            .nlargest(10)
            .sort_values()
        )
        for i, patho in enumerate(top_eff.index, start=2):
            ws_data.cell(i, 1, str(patho))
            ws_data.cell(i, 2).value = (
                f"=SUMIFS($J:$J,$H:$H,A{i},$I:$I,Analyses!$B$3)"
            )
            ws_data.cell(i, 2).number_format = "#,##0"
        n_eff = len(top_eff)

        chart_eff = BarChart()
        chart_eff.type = "bar"
        chart_eff.style = 11
        chart_eff.title = "Top 10 Pathologies (Effectifs)"
        chart_eff.height = 16
        chart_eff.width = 20
        chart_eff.legend = None
        chart_eff.add_data(
            Reference(ws_data, min_col=2, min_row=1, max_row=n_eff + 1),
            titles_from_data=True,
        )
        chart_eff.set_categories(
            Reference(ws_data, min_col=1, min_row=2, max_row=n_eff + 1)
        )
        if chart_eff.series:
            chart_eff.series[0].graphicalProperties.solidFill = "006B4F"
        ws.add_chart(chart_eff, "A6")

        # ===== Top 10 pathologies (montant) via cleanedData_Depenses (complet) =====
        cols = self.df_depenses.columns.tolist()
        l_patho = get_column_letter(cols.index("patho_niv1") + 1)
        l_mont = get_column_letter(cols.index("montant") + 1)
        l_annee = get_column_letter(cols.index("annee") + 1)
        sd = SHEET_CLEANED_DEPENSES

        top_mont = (
            dep[
                (dep["annee"] == annee_defaut)
                & (~dep["patho_niv1"].astype(str).str.lower().str.contains("total"))
            ]
            .groupby("patho_niv1")["montant"]
            .sum()
            .nlargest(10)
            .sort_values()
        )
        for i, patho in enumerate(top_mont.index, start=2):
            ws_data.cell(i, 4, str(patho))
            ws_data.cell(i, 5).value = (
                f"=SUMIFS('{sd}'!{l_mont}:{l_mont},"
                f"'{sd}'!{l_patho}:{l_patho},D{i},"
                f"'{sd}'!{l_annee}:{l_annee},Analyses!$B$3)"
            )
            ws_data.cell(i, 5).number_format = "#,##0"
        n_mont = len(top_mont)

        chart_mont = BarChart()
        chart_mont.type = "bar"
        chart_mont.style = 11
        chart_mont.title = "Top 10 Pathologies (Montant)"
        chart_mont.height = 16
        chart_mont.width = 20
        chart_mont.legend = None
        chart_mont.add_data(
            Reference(ws_data, min_col=5, min_row=1, max_row=n_mont + 1),
            titles_from_data=True,
        )
        chart_mont.set_categories(
            Reference(ws_data, min_col=4, min_row=2, max_row=n_mont + 1)
        )
        if chart_mont.series:
            chart_mont.series[0].graphicalProperties.solidFill = "4472C4"
        ws.add_chart(chart_mont, "O6")