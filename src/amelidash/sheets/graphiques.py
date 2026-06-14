"""Onglet Graphiques :
   - Effectif par Sexe x Classe d'âge (filtres Pathologie + Région + Année)
   - Taux de prévalence par Département (Top 15, filtres Pathologie + Année)
"""

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class OngletGraphiques:
    def __init__(self, wb, df, df_dep=None):
        self.wb = wb
        self.df = df.copy()
        self.df_dep = df_dep.copy() if df_dep is not None else None

    def create(self):
        for nom in ["Graphiques", "Calc", "CalcDep", "CalcRegSexe", "CalcEvol", "CalcAge",
                    "CalcMontant", "Source", "SourceDep", "SourceRegSexe", "SourceAge",
                    "SourceMontant"]:
            if nom in self.wb.sheetnames:
                self.wb.remove(self.wb[nom])

        df = self.df
        df["Effectif"] = pd.to_numeric(df["Effectif"], errors="coerce").fillna(0)
        if "Population de référence" in df.columns:
            df["Population de référence"] = pd.to_numeric(
                df["Population de référence"], errors="coerce"
            ).fillna(0)
        else:
            df["Population de référence"] = 0

        # ===================== Sources agrégées (cachées) =====================
        # Source 1 : Région / patho / âge / sexe / annee -> Effectif
        src1 = (
            df.groupby(
                ["Région", "patho_niv1", "Classe d'age", "Sexe", "annee"],
                observed=True,
            )["Effectif"]
            .sum()
            .reset_index()
        )
        ws_s1 = self.wb.create_sheet("Source")
        ws_s1.sheet_state = "hidden"
        ws_s1.append(
            ["Région", "patho_niv1", "Classe d'age", "Sexe", "annee", "Effectif"]
        )
        for r in src1.itertuples(index=False):
            ws_s1.append([r[0], r[1], r[2], r[3], int(r[4]), float(r[5])])

        # Source 2 : Département / patho / annee -> Effectif, Population
        src2 = (
            df.groupby(["Département", "patho_niv1", "annee"], observed=True)[
                ["Effectif", "Population de référence"]
            ]
            .sum()
            .reset_index()
        )
        ws_s2 = self.wb.create_sheet("SourceDep")
        ws_s2.sheet_state = "hidden"
        ws_s2.append(["Département", "patho_niv1", "annee", "Effectif", "Population"])
        for r in src2.itertuples(index=False):
            ws_s2.append([r[0], r[1], int(r[2]), float(r[3]), float(r[4])])

        # Source 3 : Région / patho / annee / sexe -> Effectif, Population
        src3 = (
            df.groupby(["Région", "patho_niv1", "annee", "Sexe"], observed=True)[
                ["Effectif", "Population de référence"]
            ]
            .sum()
            .reset_index()
        )
        ws_s3 = self.wb.create_sheet("SourceRegSexe")
        ws_s3.sheet_state = "hidden"
        ws_s3.append(
            ["Région", "patho_niv1", "annee", "Sexe", "Effectif", "Population"]
        )
        for r in src3.itertuples(index=False):
            ws_s3.append([r[0], r[1], int(r[2]), r[3], float(r[4]), float(r[5])])

        # Source 4 : Région / patho / annee / âge -> Effectif, Population
        src4 = (
            df.groupby(["Région", "patho_niv1", "annee", "Classe d'age"], observed=True)[
                ["Effectif", "Population de référence"]
            ]
            .sum()
            .reset_index()
        )
        ws_s4 = self.wb.create_sheet("SourceAge")
        ws_s4.sheet_state = "hidden"
        ws_s4.append(["Région", "patho_niv1", "annee", "Classe d'age", "Effectif", "Population"])
        for r in src4.itertuples(index=False):
            ws_s4.append([r[0], r[1], int(r[2]), r[3], float(r[4]), float(r[5])])

        # Source 5 (dépenses) : patho / annee / poste -> Montant
        postes = []
        if self.df_dep is not None and "montant" in self.df_dep.columns:
            dd = self.df_dep.copy()
            dd["montant"] = pd.to_numeric(dd["montant"], errors="coerce").fillna(0)
            src5 = (
                dd.groupby(["patho_niv1", "annee", "poste de dépense"], observed=True)[
                    "montant"
                ]
                .sum()
                .reset_index()
            )
            ws_s5 = self.wb.create_sheet("SourceMontant")
            ws_s5.sheet_state = "hidden"
            ws_s5.append(["patho_niv1", "annee", "poste", "Montant"])
            for r in src5.itertuples(index=False):
                ws_s5.append([r[0], int(r[1]), r[2], float(r[3])])
            postes = sorted(
                {
                    str(v).strip()
                    for v in dd["poste de dépense"].dropna().unique()
                    if v
                    and "total" not in str(v).lower()
                    and "soins" not in str(v).lower()
                    and str(v).strip().lower() != "nan"
                }
            )

        # ===================== Listes des filtres =====================
        def _vals(col, exclure=()):
            return sorted(
                {
                    str(v).strip()
                    for v in df[col].dropna().unique()
                    if v
                    and "total" not in str(v).lower()
                    and "tous" not in str(v).lower()
                    and not any(e in str(v).lower() for e in exclure)
                }
            )

        pathos = _vals("patho_niv1")
        regions = _vals("Région")
        ages = _vals("Classe d'age")
        sexes = _vals("Sexe", exclure=("ensemble",))
        if not sexes:
            sexes = sorted({str(v).strip() for v in df["Sexe"].dropna().unique() if v})
        annees = sorted(df["annee"].dropna().unique(), reverse=True)

        patho_def = pathos[0] if pathos else None
        annee_def = int(annees[0]) if len(annees) else None

        # ===================== Top 15 départements (prévalence, défaut) =====================
        sub = df[(df["patho_niv1"] == patho_def) & (df["annee"] == annee_def)]
        dep_def = (
            sub.groupby("Département", observed=True)[
                ["Effectif", "Population de référence"]
            ]
            .sum()
        )
        dep_def["prev"] = dep_def["Effectif"] / dep_def["Population de référence"].replace(
            0, pd.NA
        )
        top_dep = (
            dep_def["prev"].dropna().nlargest(15).sort_values().index.tolist()
        )

        # ===================== Feuilles =====================
        ws = self.wb.create_sheet("Graphiques")
        ws.sheet_view.showGridLines = False
        calc = self.wb.create_sheet("Calc")
        calc.sheet_state = "hidden"
        calcd = self.wb.create_sheet("CalcDep")
        calcd.sheet_state = "hidden"
        calcr = self.wb.create_sheet("CalcRegSexe")
        calcr.sheet_state = "hidden"
        calca = self.wb.create_sheet("CalcAge")
        calca.sheet_state = "hidden"
        calcm = self.wb.create_sheet("CalcMontant")
        calcm.sheet_state = "hidden"

        COLOR_VERT = "FF006B4F"
        COLOR_BLEU = "FF4472C4"
        COLOR_ROUGE = "C00000"
        COLOR_ROUGE_CLAIR = "F4A6A6"

        # --- Fond de couleur + couleur d'onglet (plus de titre) ---
        ws.sheet_properties.tabColor = COLOR_ROUGE
        fond = PatternFill("solid", fgColor="FCE9E9")
        for row in range(1, 72):
            for col in range(1, 21):
                ws.cell(row, col).fill = fond

        # --- Listes en colonnes cachées (évite la limite 255) ---
        L_PATHO_COL, L_REG_COL = 28, 29  # AB, AC
        for i, p in enumerate(pathos, start=2):
            ws.cell(i, L_PATHO_COL).value = p
        for i, r in enumerate(regions, start=2):
            ws.cell(i, L_REG_COL).value = r
        ws.column_dimensions[get_column_letter(L_PATHO_COL)].hidden = True
        ws.column_dimensions[get_column_letter(L_REG_COL)].hidden = True
        LP = get_column_letter(L_PATHO_COL)
        LR = get_column_letter(L_REG_COL)

        def _filtre(cell_lbl, cell_val, label, valeur):
            ws[cell_lbl] = label
            ws[cell_lbl].font = Font(bold=True, color="FFFFFF")
            ws[cell_lbl].fill = PatternFill("solid", fgColor=COLOR_BLEU)
            ws[cell_lbl].alignment = Alignment(horizontal="center")
            ws[cell_val] = valeur
            ws[cell_val].font = Font(bold=True)
            ws[cell_val].alignment = Alignment(horizontal="center")

        _filtre("A3", "B3", "Pathologie", patho_def or "")
        _filtre("A4", "B4", "Région", regions[0] if regions else "")
        _filtre("A5", "B5", "Année", annee_def)

        # Filtre pathologie : nom complet visible (renvoi à la ligne)
        ws["B3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 30

        dv_p = DataValidation(type="list", formula1=f"${LP}$2:${LP}${1 + len(pathos)}")
        dv_r = DataValidation(type="list", formula1=f"${LR}$2:${LR}${1 + len(regions)}")
        dv_y = DataValidation(
            type="list",
            formula1=f'"{",".join(str(int(a)) for a in annees)}"',
        )
        for dv, target in ((dv_p, "B3"), (dv_r, "B4"), (dv_y, "B5")):
            ws.add_data_validation(dv)
            dv.add(target)

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 50

        # ============ Tableau : Variation annuelle du nb de patients (I à M) ============
        annees_asc = sorted(int(a) for a in annees)
        from openpyxl.styles import Border, Side

        bord = Border(*([Side(style="thin", color="CCCCCC")] * 4))
        ws.merge_cells("I2:M2")
        ws["I2"] = "Variation annuelle du nombre de patients"
        ws["I2"].font = Font(bold=True, size=11, color="FFFFFF")
        ws["I2"].fill = PatternFill("solid", fgColor=COLOR_ROUGE)
        ws["I2"].alignment = Alignment(horizontal="center", vertical="center")

        entetes_var = ["Année", "Nb patient", "N-1 Nb patient", "Évolution %"]
        for j, txt in enumerate(entetes_var, start=9):
            c = ws.cell(3, j, txt)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=COLOR_BLEU)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bord

        for k, an in enumerate(annees_asc):
            r = 4 + k
            ws.cell(r, 9, an).border = bord
            ws.cell(r, 9).alignment = Alignment(horizontal="center")
            ws.cell(r, 10).value = (
                f"=SUMIFS(SourceRegSexe!$E:$E,"
                f"SourceRegSexe!$A:$A,Graphiques!$B$4,"
                f"SourceRegSexe!$B:$B,Graphiques!$B$3,"
                f"SourceRegSexe!$C:$C,I{r})"
            )
            ws.cell(r, 10).number_format = "#,##0"
            ws.cell(r, 10).border = bord
            if k == 0:
                ws.cell(r, 11).value = 0
                ws.cell(r, 12).value = 0
            else:
                ws.cell(r, 11).value = f"=J{r - 1}"
                ws.cell(r, 12).value = f"=IFERROR((J{r}-K{r})/K{r},0)"
            ws.cell(r, 11).number_format = "#,##0"
            ws.cell(r, 11).border = bord
            ws.cell(r, 12).number_format = "0.00%"
            ws.cell(r, 12).border = bord

        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 16
        ws.column_dimensions["L"].width = 14

        # ============ Matrice 1 : Effectif (Sexe x Âge) ============
        calc.cell(1, 1, "Sexe")
        for j, age in enumerate(ages, start=2):
            calc.cell(1, j, age)
        for i, sexe in enumerate(sexes, start=2):
            calc.cell(i, 1, sexe)
            for j in range(2, 2 + len(ages)):
                col_age = get_column_letter(j)
                calc.cell(i, j).value = (
                    f"=SUMIFS(Source!$F:$F,"
                    f"Source!$A:$A,Graphiques!$B$4,"
                    f"Source!$B:$B,Graphiques!$B$3,"
                    f"Source!$C:$C,{col_age}$1,"
                    f"Source!$D:$D,$A{i},"
                    f"Source!$E:$E,Graphiques!$B$5)"
                )
                calc.cell(i, j).number_format = "#,##0"
        n_sexe, n_age = len(sexes), len(ages)

        chart1 = BarChart()
        chart1.type = "col"
        chart1.grouping = "clustered"
        chart1.title = None
        chart1.height = 11
        chart1.width = 16
        chart1.style = 12
        chart1.add_data(
            Reference(calc, min_col=2, max_col=1 + n_age, min_row=1, max_row=1 + n_sexe),
            titles_from_data=True,
        )
        chart1.set_categories(Reference(calc, min_col=1, min_row=2, max_row=1 + n_sexe))
        chart1.x_axis.title = "Sexe"
        chart1.y_axis.title = "Effectif"
        chart1.x_axis.delete = False
        chart1.y_axis.delete = False
        chart1.legend.position = "t"
        ws.add_chart(chart1, "A8")

        # ============ Matrice 2 : Prévalence par Département (Top 15) ============
        calcd.cell(1, 1, "Département")
        calcd.cell(1, 2, "Effectif")
        calcd.cell(1, 3, "Population")
        calcd.cell(1, 4, "Prévalence")
        for i, dep in enumerate(top_dep, start=2):
            calcd.cell(i, 1, dep)
            calcd.cell(i, 2).value = (
                f"=SUMIFS(SourceDep!$D:$D,SourceDep!$A:$A,A{i},"
                f"SourceDep!$B:$B,Graphiques!$B$3,SourceDep!$C:$C,Graphiques!$B$5)"
            )
            calcd.cell(i, 3).value = (
                f"=SUMIFS(SourceDep!$E:$E,SourceDep!$A:$A,A{i},"
                f"SourceDep!$B:$B,Graphiques!$B$3,SourceDep!$C:$C,Graphiques!$B$5)"
            )
            calcd.cell(i, 4).value = f"=IF(C{i}=0,0,B{i}/C{i})"
            calcd.cell(i, 4).number_format = "0.00%"
        n_dep = len(top_dep)

        chart2 = BarChart()
        chart2.type = "bar"  # barres horizontales
        chart2.title = "Top 15 — Taux de prévalence par Département"
        chart2.height = 12
        chart2.width = 16
        chart2.style = 10
        chart2.legend = None
        chart2.add_data(
            Reference(calcd, min_col=4, min_row=1, max_row=n_dep + 1),
            titles_from_data=True,
        )
        chart2.set_categories(Reference(calcd, min_col=1, min_row=2, max_row=n_dep + 1))
        chart2.x_axis.title = "Département"
        chart2.y_axis.title = None
        chart2.x_axis.delete = False
        chart2.y_axis.delete = False
        if chart2.series:
            chart2.series[0].graphicalProperties.solidFill = COLOR_ROUGE
        ws.add_chart(chart2, "I8")

        # ============ Matrice 3 : Prévalence par Région, Homme vs Femme ============
        calcr.cell(1, 1, "Région")
        for j, sexe in enumerate(sexes, start=2):
            calcr.cell(1, j, sexe)
        for i, region in enumerate(regions, start=2):
            calcr.cell(i, 1, region)
            for j in range(2, 2 + len(sexes)):
                col_sexe = get_column_letter(j)
                calcr.cell(i, j).value = (
                    f"=IFERROR("
                    f"SUMIFS(SourceRegSexe!$E:$E,SourceRegSexe!$A:$A,$A{i},"
                    f"SourceRegSexe!$B:$B,Graphiques!$B$3,SourceRegSexe!$C:$C,Graphiques!$B$5,"
                    f"SourceRegSexe!$D:$D,{col_sexe}$1)"
                    f"/"
                    f"SUMIFS(SourceRegSexe!$F:$F,SourceRegSexe!$A:$A,$A{i},"
                    f"SourceRegSexe!$B:$B,Graphiques!$B$3,SourceRegSexe!$C:$C,Graphiques!$B$5,"
                    f"SourceRegSexe!$D:$D,{col_sexe}$1)"
                    f",0)"
                )
                calcr.cell(i, j).number_format = "0.00%"
        n_reg = len(regions)

        chart3 = BarChart()
        chart3.type = "col"
        chart3.grouping = "clustered"
        chart3.title = None
        chart3.height = 11
        chart3.width = 16
        chart3.style = 12
        chart3.add_data(
            Reference(calcr, min_col=2, max_col=1 + len(sexes), min_row=1, max_row=1 + n_reg),
            titles_from_data=True,
        )
        chart3.set_categories(Reference(calcr, min_col=1, min_row=2, max_row=1 + n_reg))
        chart3.x_axis.title = "Région"
        chart3.y_axis.title = "Taux de prévalence"
        chart3.x_axis.delete = False
        chart3.y_axis.delete = False
        chart3.legend.position = "t"
        rouges = [COLOR_ROUGE, COLOR_ROUGE_CLAIR]
        for idx, serie in enumerate(chart3.series):
            serie.graphicalProperties.solidFill = rouges[idx % len(rouges)]
        ws.add_chart(chart3, "A32")

        # ============ Matrice 5 : Prévalence par classe d'âge ============
        calca.cell(1, 1, "Classe d'âge")
        calca.cell(1, 2, "Prévalence")
        for i, age in enumerate(ages, start=2):
            calca.cell(i, 1, age)
            calca.cell(i, 2).value = (
                f"=IFERROR("
                f"SUMIFS(SourceAge!$E:$E,SourceAge!$A:$A,Graphiques!$B$4,"
                f"SourceAge!$B:$B,Graphiques!$B$3,SourceAge!$C:$C,Graphiques!$B$5,"
                f"SourceAge!$D:$D,A{i})"
                f"/"
                f"SUMIFS(SourceAge!$F:$F,SourceAge!$A:$A,Graphiques!$B$4,"
                f"SourceAge!$B:$B,Graphiques!$B$3,SourceAge!$C:$C,Graphiques!$B$5,"
                f"SourceAge!$D:$D,A{i})"
                f",0)"
            )
            calca.cell(i, 2).number_format = "0.00%"
        n_age2 = len(ages)

        chart5 = BarChart()
        chart5.type = "col"
        chart5.title = "Taux de prévalence par classe d'âge"
        chart5.height = 11
        chart5.width = 16
        chart5.style = 10
        chart5.legend = None
        chart5.add_data(
            Reference(calca, min_col=2, min_row=1, max_row=n_age2 + 1),
            titles_from_data=True,
        )
        chart5.set_categories(Reference(calca, min_col=1, min_row=2, max_row=n_age2 + 1))
        chart5.x_axis.title = "Classe d'âge"
        chart5.y_axis.title = "Taux de prévalence"
        chart5.x_axis.delete = False
        chart5.y_axis.delete = False
        if chart5.series:
            chart5.series[0].graphicalProperties.solidFill = COLOR_ROUGE
        ws.add_chart(chart5, "I32")

        # ============ Matrice 6 : Montant par poste de dépense (si dépenses fournies) ============
        if postes:
            calcm.cell(1, 1, "Poste")
            calcm.cell(1, 2, "Montant")
            for i, poste in enumerate(postes, start=2):
                calcm.cell(i, 1, poste)
                calcm.cell(i, 2).value = (
                    f"=SUMIFS(SourceMontant!$D:$D,"
                    f"SourceMontant!$A:$A,Graphiques!$B$3,"
                    f"SourceMontant!$B:$B,Graphiques!$B$5,"
                    f"SourceMontant!$C:$C,A{i})"
                )
                calcm.cell(i, 2).number_format = "#,##0 €"
            n_post = len(postes)

            chart6 = BarChart()
            chart6.type = "col"
            chart6.title = "Montant par poste de dépense (€)"
            chart6.height = 11
            chart6.width = 16
            chart6.style = 12
            chart6.legend = None
            chart6.add_data(
                Reference(calcm, min_col=2, min_row=1, max_row=n_post + 1),
                titles_from_data=True,
            )
            chart6.set_categories(
                Reference(calcm, min_col=1, min_row=2, max_row=n_post + 1)
            )
            chart6.x_axis.title = "Poste"
            chart6.y_axis.title = "Montant (€)"
            chart6.x_axis.delete = False
            chart6.y_axis.delete = False
            if chart6.series:
                chart6.series[0].graphicalProperties.solidFill = COLOR_ROUGE
            ws.add_chart(chart6, "A56")