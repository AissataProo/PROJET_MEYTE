import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference, LineChart, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from config import SHEET_CLEANED_EFFECTIFS


class OngletDepartement:
    """Crée l'onglet Departement avec filtres dynamiques et graphiques dynamiques."""

    def __init__(self, wb, df, df_dep=None):
        self.wb = wb
        self.df = df.copy()
        self.df_dep = df_dep.copy() if df_dep is not None else None
        self.sheet_eff = SHEET_CLEANED_EFFECTIFS

    def _get_column_mapping(self):
        """Récupère le mapping colonnes de la feuille nettoyée."""
        ws_eff = self.wb[self.sheet_eff]
        hdr = {
            str(c.value).strip(): get_column_letter(i)
            for i, c in enumerate(ws_eff[1], start=1)
            if c.value
        }
        return {
            "annee": hdr.get("annee"),
            "sexe": hdr.get("Sexe"),
            "patho": hdr.get("patho_niv1"),
            "age": hdr.get("Classe d'age"),
            "dept": hdr.get("Département"),
            "region": hdr.get("Région"),
            "eff": hdr.get("Effectif"),
            "prevalence": hdr.get("Prévalence"),
            "pop_ref": hdr.get("Population de référence"),
        }

    def _add_filter(self, ws, row, label, values, ws_listes, col_idx):
        """Ajoute un filtre. La liste est stockée dans une feuille cachée
        (pas de limite 255 car., et pas de showDropDown qui casse la flèche)."""
        COLOR_BLEU = "FF4472C4"
        COLOR_BLANC = "FFFFFFFF"

        # Label
        ws.cell(row, 1).value = label
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=COLOR_BLEU)
        ws.cell(row, 1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")

        # Valeur avec dropdown
        default = values[0] if values else ""
        ws.cell(row, 2).value = default
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=COLOR_BLANC)
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")

        # Écriture des valeurs dans la feuille cachée (1 colonne par filtre)
        col_letter = get_column_letter(col_idx)
        for i, v in enumerate(values, 1):
            ws_listes.cell(i, col_idx, v)
        last = max(len(values), 1)

        # Validation pointant vers la plage
        plage = f"'{ws_listes.title}'!${col_letter}$1:${col_letter}${last}"
        dv = DataValidation(type="list", formula1=plage, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws.cell(row, 2).coordinate)

    def _create_chart(self, chart_type, title):
        """Crée et formate un graphique."""
        if chart_type == "BarChart":
            chart = BarChart()
        elif chart_type == "LineChart":
            chart = LineChart()
            chart.smooth = True
        else:
            chart = PieChart()

        chart.title = title
        chart.style = 12
        chart.legend.position = "r"
        return chart

    def create(self):
        """Crée l'onglet Departement avec 2 filtres (Département, Année) et graphiques."""

        # Préparation des données
        df = self.df.copy()
        df["Effectif"] = pd.to_numeric(df["Effectif"], errors="coerce").fillna(0)

        if "Population de référence" in df.columns:
            df["Population de référence"] = pd.to_numeric(
                df["Population de référence"], errors="coerce"
            ).fillna(0)

        # Créer les listes de valeurs uniques
        deps = sorted(df["Département"].dropna().unique())
        annees = sorted([int(a) for a in df["annee"].dropna().unique()], reverse=True)
        regions = sorted(df["Région"].dropna().unique())
        pathos = sorted(df["patho_niv1"].dropna().unique())
        sexes = sorted(
            [
                s
                for s in df["Sexe"].dropna().unique()
                if "ensemble" not in str(s).lower()
            ]
        )
        ages = sorted(df["Classe d'age"].dropna().unique())

        # 1. Nettoyage des anciennes feuilles
        for nom in [
            "Departement",
            "CalcListes",
            "CalcTopPatho",
            "CalcPathoSexe",
            "CalcRegion",
            "CalcAgeEff",
            "CalcSexe",
            "CalcPrevAge",
            "CalcVariation",
            "CalcCorrelation",
        ]:
            if nom in self.wb.sheetnames:
                self.wb.remove(self.wb[nom])

        # 2. Création de l'onglet
        ws = self.wb.create_sheet("Departement")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A8"

        col_map = self._get_column_mapping()
        COLOR_BLEU = "FF4472C4"

        # Feuille cachée qui stocke les listes des filtres
        calc_listes = self.wb.create_sheet("CalcListes")
        calc_listes.sheet_state = "hidden"

        # 3. FILTRES (Département + Année uniquement)
        self._add_filter(ws, 3, "Département", list(deps), calc_listes, 1)
        self._add_filter(ws, 4, "Année", list(annees), calc_listes, 2)

        # 4. INDICATEURS CLÉS
        ws.merge_cells("E3:H3")
        ws["E3"] = "Indicateurs clés"
        ws["E3"].fill = PatternFill("solid", fgColor=COLOR_BLEU)
        ws["E3"].font = Font(bold=True, color="FFFFFF")

        ws["E4"] = "Total patients :"
        ws["F4"].value = (
            f"=SUMIFS('{self.sheet_eff}'!{col_map['eff']}:{col_map['eff']},"
            f"'{self.sheet_eff}'!{col_map['dept']}:{col_map['dept']},B3,"
            f"'{self.sheet_eff}'!{col_map['annee']}:{col_map['annee']},B4)"
        )
        ws["F4"].number_format = "#,##0"

        # 5. FEUILLES DE CALCUL CACHÉES
        calc_top_patho = self.wb.create_sheet("CalcTopPatho")
        calc_top_patho.sheet_state = "hidden"

        calc_patho_sexe = self.wb.create_sheet("CalcPathoSexe")
        calc_patho_sexe.sheet_state = "hidden"

        calc_region = self.wb.create_sheet("CalcRegion")
        calc_region.sheet_state = "hidden"

        calc_age_eff = self.wb.create_sheet("CalcAgeEff")
        calc_age_eff.sheet_state = "hidden"

        calc_prev_age = self.wb.create_sheet("CalcPrevAge")
        calc_prev_age.sheet_state = "hidden"

        # 6. GRAPHIQUE 1: Top pathologies par effectif (filtré Département + Année)
        calc_top_patho["A1"] = "Pathologie"
        calc_top_patho["B1"] = "Effectif"

        for i, patho in enumerate(pathos, 2):
            calc_top_patho.cell(i, 1, patho)
            calc_top_patho.cell(i, 2).value = (
                f"=SUMIFS('{self.sheet_eff}'!{col_map['eff']}:{col_map['eff']},"
                f"'{self.sheet_eff}'!{col_map['patho']}:{col_map['patho']},A{i},"
                f"'{self.sheet_eff}'!{col_map['dept']}:{col_map['dept']},Departement!$B$3,"
                f"'{self.sheet_eff}'!{col_map['annee']}:{col_map['annee']},Departement!$B$4)"
            )

        chart1 = self._create_chart("BarChart", "Top pathologies par effectif")
        chart1.type = "bar"
        chart1.add_data(
            Reference(calc_top_patho, min_col=2, min_row=1, max_row=len(pathos) + 1),
            titles_from_data=True,
        )
        chart1.set_categories(
            Reference(calc_top_patho, min_col=1, min_row=2, max_row=len(pathos) + 1)
        )
        chart1.height, chart1.width = 12, 16
        ws.add_chart(chart1, "A8")

        # 7. GRAPHIQUE 2: Effectif par pathologie et sexe (filtré Département + Année)
        calc_patho_sexe["A1"] = "Pathologie"
        for i, s in enumerate(sexes, 2):
            calc_patho_sexe.cell(1, i, s)

        for r, p in enumerate(pathos, 2):
            calc_patho_sexe.cell(r, 1, p)
            for c, s in enumerate(sexes, 2):
                calc_patho_sexe.cell(r, c).value = (
                    f"=SUMIFS('{self.sheet_eff}'!{col_map['eff']}:{col_map['eff']},"
                    f"'{self.sheet_eff}'!{col_map['patho']}:{col_map['patho']},$A{r},"
                    f"'{self.sheet_eff}'!{col_map['sexe']}:{col_map['sexe']},'{calc_patho_sexe.title}'!${get_column_letter(c)}$1,"
                    f"'{self.sheet_eff}'!{col_map['dept']}:{col_map['dept']},Departement!$B$3,"
                    f"'{self.sheet_eff}'!{col_map['annee']}:{col_map['annee']},Departement!$B$4)"
                )

        chart2 = self._create_chart("BarChart", "Effectif par pathologie et sexe")
        chart2.add_data(
            Reference(
                calc_patho_sexe,
                min_col=2,
                max_col=1 + len(sexes),
                min_row=1,
                max_row=len(pathos) + 1,
            ),
            titles_from_data=True,
        )
        chart2.set_categories(
            Reference(calc_patho_sexe, min_col=1, min_row=2, max_row=len(pathos) + 1)
        )
        chart2.height, chart2.width = 12, 18
        ws.add_chart(chart2, "K8")

        # Effectif par région
        calc_region["A1"] = "Région"
        calc_region["B1"] = "Effectif"

        for i, region in enumerate(regions, 2):
            calc_region.cell(i, 1, region)
            calc_region.cell(i, 2).value = (
                f"=SUMIFS('{self.sheet_eff}'!{col_map['eff']}:{col_map['eff']},"
                f"'{self.sheet_eff}'!{col_map['region']}:{col_map['region']},A{i},"
                f"'{self.sheet_eff}'!{col_map['annee']}:{col_map['annee']},Departement!$B$4)"
            )

        chart3 = self._create_chart("BarChart", "Effectif par région")
        chart3.type = "col"
        chart3.add_data(
            Reference(calc_region, min_col=2, min_row=1, max_row=len(regions) + 1),
            titles_from_data=True,
        )
        chart3.set_categories(
            Reference(calc_region, min_col=1, min_row=2, max_row=len(regions) + 1)
        )
        chart3.height, chart3.width = 10, 16
        ws.add_chart(chart3, "A33")

        # Effectif par classe d'âge
        calc_age_eff["A1"] = "Classe d'âge"
        calc_age_eff["B1"] = "Effectif"

        for i, age in enumerate(ages, 2):
            calc_age_eff.cell(i, 1, age)
            calc_age_eff.cell(i, 2).value = (
                f"=SUMIFS('{self.sheet_eff}'!{col_map['eff']}:{col_map['eff']},"
                f"'{self.sheet_eff}'!{col_map['age']}:{col_map['age']},A{i},"
                f"'{self.sheet_eff}'!{col_map['dept']}:{col_map['dept']},Departement!$B$3,"
                f"'{self.sheet_eff}'!{col_map['annee']}:{col_map['annee']},Departement!$B$4)"
            )

        chart4 = self._create_chart("BarChart", "Effectif par classe d'âge")
        chart4.add_data(
            Reference(calc_age_eff, min_col=2, min_row=1, max_row=len(ages) + 1),
            titles_from_data=True,
        )
        chart4.set_categories(
            Reference(calc_age_eff, min_col=1, min_row=2, max_row=len(ages) + 1)
        )
        chart4.height, chart4.width = 10, 16
        ws.add_chart(chart4, "K33")

        # Prévalence par classe d'âge (filtré Département + Année)
        calc_prev_age["A1"] = "Classe d'âge"
        calc_prev_age["B1"] = "Prévalence (%)"

        for i, age in enumerate(ages, 2):
            calc_prev_age.cell(i, 1, age)
            if col_map["prevalence"]:
                calc_prev_age.cell(i, 2).value = (
                    f"=IFERROR(SUMIFS('{self.sheet_eff}'!{col_map['prevalence']}:{col_map['prevalence']},"
                    f"'{self.sheet_eff}'!{col_map['age']}:{col_map['age']},A{i},"
                    f"'{self.sheet_eff}'!{col_map['dept']}:{col_map['dept']},Departement!$B$3,"
                    f"'{self.sheet_eff}'!{col_map['annee']}:{col_map['annee']},Departement!$B$4)"
                    f"/COUNTIFS('{self.sheet_eff}'!{col_map['age']}:{col_map['age']},A{i},"
                    f"'{self.sheet_eff}'!{col_map['dept']}:{col_map['dept']},Departement!$B$3,"
                    f"'{self.sheet_eff}'!{col_map['annee']}:{col_map['annee']},Departement!$B$4),0)"
                )

        chart5 = self._create_chart("LineChart", "Prévalence par classe d'âge")
        chart5.add_data(
            Reference(calc_prev_age, min_col=2, min_row=1, max_row=len(ages) + 1),
            titles_from_data=True,
        )
        chart5.set_categories(
            Reference(calc_prev_age, min_col=1, min_row=2, max_row=len(ages) + 1)
        )
        chart5.height, chart5.width = 10, 16
        ws.add_chart(chart5, "A56")

        # Corrélation Population / Effectif (Scatter, 1 série = 1 pathologie)
        if col_map.get("pop_ref"):
            calc_corr = self.wb.create_sheet("CalcCorrelation")
            calc_corr.sheet_state = "hidden"
            calc_corr["A1"], calc_corr["B1"], calc_corr["C1"] = (
                "Pathologie",
                "Population de référence",
                "Effectif",
            )

            for i, patho in enumerate(pathos, 2):
                calc_corr.cell(i, 1, patho)
                calc_corr.cell(i, 2).value = (
                    f"=SUMIFS('{self.sheet_eff}'!{col_map['pop_ref']}:{col_map['pop_ref']},"
                    f"'{self.sheet_eff}'!{col_map['patho']}:{col_map['patho']},A{i},"
                    f"'{self.sheet_eff}'!{col_map['dept']}:{col_map['dept']},Departement!$B$3,"
                    f"'{self.sheet_eff}'!{col_map['annee']}:{col_map['annee']},Departement!$B$4)"
                )
                calc_corr.cell(i, 3).value = (
                    f"=SUMIFS('{self.sheet_eff}'!{col_map['eff']}:{col_map['eff']},"
                    f"'{self.sheet_eff}'!{col_map['patho']}:{col_map['patho']},A{i},"
                    f"'{self.sheet_eff}'!{col_map['dept']}:{col_map['dept']},Departement!$B$3,"
                    f"'{self.sheet_eff}'!{col_map['annee']}:{col_map['annee']},Departement!$B$4)"
                )

            scatter = ScatterChart()
            scatter.title = "Corrélation Population / Effectif"
            scatter.style = 13
            scatter.x_axis.title = "Population de référence"
            scatter.y_axis.title = "Effectif"
            scatter.legend.position = "r" 

            for i, patho in enumerate(pathos, 2):
                xref = Reference(calc_corr, min_col=2, min_row=i, max_row=i)
                yref = Reference(calc_corr, min_col=3, min_row=i, max_row=i)
                s = Series(yref, xref, title=str(patho))
                s.marker = Marker(symbol="circle", size=7)
                s.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
                scatter.series.append(s)

            scatter.height, scatter.width = 10, 16
            ws.add_chart(scatter, "K56")
            # --- TABLEAU : Variation annuelle du nombre de patients (J → N) ---

            ROUGE = "FFC00000"
            thin = Side(style="thin", color="FFBFBFBF")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            start_row = 3      # ligne 3
            start_col = 10     # colonne J

            # Titre
            ws.merge_cells(start_row=start_row, start_column=start_col,
                        end_row=start_row, end_column=start_col+4)
            tcell = ws.cell(start_row, start_col, "Variation annuelle du nombre de patients")
            tcell.fill = PatternFill("solid", fgColor=ROUGE)
            tcell.font = Font(bold=True, color="FFFFFF")
            tcell.alignment = Alignment(horizontal="left", vertical="center")

            # En-têtes
            headers = ["Année", "Nb patients", "Nb patients N-1", "Var %"]
            for i, txt in enumerate(headers):
                cell = ws.cell(start_row + 1, start_col + i, txt)
                cell.fill = PatternFill("solid", fgColor=COLOR_BLEU)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # --- BOUCLE SUR LES ANNÉES (OBLIGATOIRE) ---
            annees_asc = sorted(annees)

            for k, an in enumerate(annees_asc):
                r = start_row + 2 + k 

                # Année (colonne J)
                ws.cell(r, start_col, an).border = border

                # Nb patients N
                ws.cell(
                    r, start_col + 1,
                    f"=SOMME.SI.ENS(cleanedData_Effectifs!G:G;"
                    f"cleanedData_Effectifs!A:A;Departement!J{r};"
                    f"cleanedData_Effectifs!F:F;Departement!$B$3)"
                ).number_format = "#,##0"
                ws.cell(r, start_col + 1).border = border

                # Nb patients N-1
                ws.cell(
                    r, start_col + 2,
                    f"=SOMME.SI.ENS(cleanedData_Effectifs!G:G;"
                    f"cleanedData_Effectifs!A:A;Departement!J{r}-1;"
                    f"cleanedData_Effectifs!F:F;Departement!$B$3)"
                ).number_format = "#,##0"
                ws.cell(r, start_col + 2).border = border

                # Var %
                ws.cell(
                    r, start_col + 3,
                    f"=SIERREUR((K{r}-L{r})/L{r};\"\")"
                ).number_format = "0.0%"
                ws.cell(r, start_col + 3).alignment = Alignment(horizontal="center")
                ws.cell(r, start_col + 3).border = border