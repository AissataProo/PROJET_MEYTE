"""Onglet Postes : tableau de bord depenses par poste avec filtres et graphique pie."""

from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from config import SHEET_CLEANED_DEPENSES


class OngletPostes:
    def __init__(self, wb, df):
        self.wb = wb
        self.df = df
        self.sheet_cleaned_dep = SHEET_CLEANED_DEPENSES
        self.create()

    def create(self):
        ws = self.wb.create_sheet("Postes")
        ws.sheet_view.showGridLines = False

        ws_dep = self.wb[self.sheet_cleaned_dep]
        sheet_dep = self.sheet_cleaned_dep

        headers = {}
        for idx, cell in enumerate(ws_dep[1], start=1):
            if cell.value is not None:
                headers[str(cell.value).strip()] = idx

        col_annee = headers.get("annee", 1)
        col_poste = headers.get("poste de dépense", 5)
        col_patho3 = headers.get("patho_niv3", 4)
        col_montant = headers.get("montant", 7)
        col_effectif = headers.get("Effectif", 8)

        col_annee_letter = get_column_letter(col_annee)
        col_poste_letter = get_column_letter(col_poste)
        col_patho3_letter = get_column_letter(col_patho3)
        col_montant_letter = get_column_letter(col_montant)
        col_effectif_letter = get_column_letter(col_effectif)

        annees_list = sorted(self.df["annee"].dropna().unique(), reverse=True)
        annee_sel = annees_list[0] if len(annees_list) else None

        postes_base = sorted(
            {
                str(v).strip()
                for v in self.df["poste de dépense"].dropna().unique()
                if v and "total" not in str(v).lower()
            }
        )

        postes_with_amount = []
        for poste in postes_base:
            montant = 0
            for row in ws_dep.iter_rows(
                min_row=2, max_row=ws_dep.max_row, values_only=True
            ):
                sous_poste = (
                    str(row[col_patho3 - 1]).strip() if row[col_patho3 - 1] else ""
                )

                if (
                    row[col_poste - 1]
                    and str(row[col_poste - 1]).strip() == poste
                    and str(row[col_annee - 1]).strip() == str(annee_sel)
                    and "total" not in sous_poste.lower()
                    and "soins" not in sous_poste.lower()
                    and sous_poste.lower() != "nan"
                ):
                    try:
                        montant += float(row[col_montant - 1] or 0)
                    except Exception:
                        pass
            postes_with_amount.append((poste, montant))

        postes_with_amount.sort(key=lambda x: x[1], reverse=True)
        postes_list = [p for p, _ in postes_with_amount]

        poste_defaut = postes_list[0] if postes_list else "Sélectionner"

        patho3_list = sorted(
            set(
                str(row[col_patho3 - 1]).strip()
                for row in ws_dep.iter_rows(min_row=2, values_only=True)
                if row[col_patho3 - 1]
                and str(row[col_poste - 1]).strip() == poste_defaut
                and "total" not in str(row[col_patho3 - 1]).strip().lower()
                and "soins" not in str(row[col_patho3 - 1]).strip().lower()
                and str(row[col_patho3 - 1]).strip().lower() != "nan"
            )
        )

        COLOR_VERT = "FF006B4F"
        COLOR_BLEU = "FF4472C4"

        ws.merge_cells("B1:L1")
        ws["B1"] = "Analyse des Pathologies"
        ws["B1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["B1"].fill = PatternFill(start_color=COLOR_VERT, fill_type="solid")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws["A3"] = "Année"
        ws["A3"].font = Font(bold=True, color="FFFFFF")
        ws["A3"].fill = PatternFill(start_color=COLOR_BLEU, fill_type="solid")
        ws["A3"].alignment = Alignment(horizontal="center")

        ws["B3"] = annees_list[0] if annees_list else None
        ws["B3"].fill = PatternFill(start_color="FFFFFFFF", fill_type="solid")
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].alignment = Alignment(horizontal="center")

        annees_str = ",".join(str(int(a)) for a in annees_list)
        dv_annee = DataValidation(
            type="list", formula1=f'"{annees_str}"', allow_blank=False
        )
        ws.add_data_validation(dv_annee)
        dv_annee.add("B3")

        ws["A5"] = "Poste"
        ws["A5"].font = Font(bold=True, color="FFFFFF")
        ws["A5"].fill = PatternFill(start_color=COLOR_BLEU, fill_type="solid")
        ws["A5"].alignment = Alignment(horizontal="center")

        ws["B5"] = poste_defaut
        ws["B5"].fill = PatternFill(start_color="FFFFFFFF", fill_type="solid")
        ws["B5"].font = Font(bold=True, size=11)
        ws["B5"].alignment = Alignment(horizontal="center")

        postes_str = ",".join(postes_list)
        dv_poste = DataValidation(
            type="list", formula1=f'"{postes_str}"', allow_blank=False
        )
        ws.add_data_validation(dv_poste)
        dv_poste.add("B5")

        ws.merge_cells("A8:B8")
        ws["A8"] = "Total des dépenses"
        ws["A8"].font = Font(bold=True, size=11, color="FFFFFF")
        ws["A8"].fill = PatternFill(start_color=COLOR_VERT, fill_type="solid")
        ws["A8"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[8].height = 22

        ws["C8"] = (
            f"=IFERROR(SUMIFS('{sheet_dep}'!${col_montant_letter}$2:${col_montant_letter}$5000,"
            f"'{sheet_dep}'!${col_annee_letter}$2:${col_annee_letter}$5000,B$3,"
            f"'{sheet_dep}'!${col_poste_letter}$2:${col_poste_letter}$5000,B$5),0)"
        )
        ws["C8"].font = Font(bold=True, size=13, color="FF006B4F")
        ws["C8"].number_format = "#,##0 €"
        ws["C8"].alignment = Alignment(horizontal="left")

        border_header = Border(
            left=Side(style="thin", color="4472C4"),
            right=Side(style="thin", color="4472C4"),
            top=Side(style="thin", color="4472C4"),
            bottom=Side(style="thin", color="4472C4"),
        )
        border_data = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        couleur_alt_1 = PatternFill(start_color="FFF5F5F5", fill_type="solid")
        couleur_alt_2 = PatternFill(start_color="FFFFFFFF", fill_type="solid")

        entetes = ["Pathologie", "Montant", "Effectif", "Coût/patient"]
        for i, texte in enumerate(entetes, start=1):
            cell = ws.cell(10, i)
            cell.value = texte
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=COLOR_BLEU, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_header
        ws.row_dimensions[10].height = 22

        for idx, patho3 in enumerate(patho3_list):
            row = 12 + idx
            couleur = couleur_alt_1 if idx % 2 == 0 else couleur_alt_2

            ws[f"A{row}"] = patho3
            ws[f"A{row}"].fill = couleur
            ws[f"A{row}"].border = border_data
            ws[f"A{row}"].alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            ws[f"A{row}"].font = Font(size=10)

            ws[f"B{row}"] = (
                f"=IFERROR(SUMIFS('{sheet_dep}'!${col_montant_letter}$2:${col_montant_letter}$5000,"
                f"'{sheet_dep}'!${col_annee_letter}$2:${col_annee_letter}$5000,B$3,"
                f"'{sheet_dep}'!${col_poste_letter}$2:${col_poste_letter}$5000,B$5,"
                f"'{sheet_dep}'!${col_patho3_letter}$2:${col_patho3_letter}$5000,A{row}),0)"
            )
            ws[f"B{row}"].number_format = "#,##0"
            ws[f"B{row}"].fill = couleur
            ws[f"B{row}"].border = border_data
            ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")

            ws[f"C{row}"] = (
                f"=IFERROR(SUMIFS('{sheet_dep}'!${col_effectif_letter}$2:${col_effectif_letter}$5000,"
                f"'{sheet_dep}'!${col_annee_letter}$2:${col_annee_letter}$5000,B$3,"
                f"'{sheet_dep}'!${col_poste_letter}$2:${col_poste_letter}$5000,B$5,"
                f"'{sheet_dep}'!${col_patho3_letter}$2:${col_patho3_letter}$5000,A{row}),0)"
            )
            ws[f"C{row}"].number_format = "#,##0"
            ws[f"C{row}"].fill = couleur
            ws[f"C{row}"].border = border_data
            ws[f"C{row}"].alignment = Alignment(horizontal="right", vertical="center")

            ws[f"D{row}"] = f"=IFERROR(B{row}/C{row},0)"
            ws[f"D{row}"].number_format = "#,##0.00 €"
            ws[f"D{row}"].fill = couleur
            ws[f"D{row}"].border = border_data
            ws[f"D{row}"].alignment = Alignment(horizontal="right", vertical="center")

            ws.row_dimensions[row].height = 25

        last_row = 12 + len(patho3_list) - 1

        S_COMP = last_row + 3

        ws.merge_cells(f"A{S_COMP}:E{S_COMP}")
        ws[f"A{S_COMP}"] = "Comparaison annuelle"
        ws[f"A{S_COMP}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{S_COMP}"].fill = PatternFill(start_color=COLOR_VERT, fill_type="solid")
        ws[f"A{S_COMP}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[S_COMP].height = 25

        COMP_HDR = S_COMP + 1
        COMP_START = COMP_HDR + 1

        ws.cell(COMP_HDR, 1).value = "Pathologie"
        ws.cell(COMP_HDR, 2).value = "Dép. 2022"
        ws.cell(COMP_HDR, 3).value = "Dép. 2023"
        ws.cell(COMP_HDR, 4).value = "Variation"
        ws.cell(COMP_HDR, 5).value = "Évol. %"
        for cell in [ws.cell(COMP_HDR, i) for i in range(1, 6)]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=COLOR_BLEU, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_header
        ws.row_dimensions[COMP_HDR].height = 22

        for idx, patho in enumerate(patho3_list):
            r = COMP_START + idx
            couleur = couleur_alt_1 if idx % 2 == 0 else couleur_alt_2

            ws.cell(r, 1).value = patho
            ws.cell(r, 1).fill = couleur
            ws.cell(r, 1).border = border_data
            ws.cell(r, 1).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            ws.cell(r, 1).font = Font(size=10)

            ws.cell(r, 2).value = (
                f"=IFERROR(SUMIFS('{sheet_dep}'!${col_montant_letter}:${col_montant_letter},"
                f"'{sheet_dep}'!${col_patho3_letter}:${col_patho3_letter},A{r},"
                f"'{sheet_dep}'!${col_poste_letter}:${col_poste_letter},B$5,"
                f"'{sheet_dep}'!${col_annee_letter}:${col_annee_letter},2022),0)"
            )
            ws.cell(r, 2).number_format = "#,##0"
            ws.cell(r, 2).fill = couleur
            ws.cell(r, 2).border = border_data
            ws.cell(r, 2).alignment = Alignment(horizontal="right")

            ws.cell(r, 3).value = (
                f"=IFERROR(SUMIFS('{sheet_dep}'!${col_montant_letter}:${col_montant_letter},"
                f"'{sheet_dep}'!${col_patho3_letter}:${col_patho3_letter},A{r},"
                f"'{sheet_dep}'!${col_poste_letter}:${col_poste_letter},B$5,"
                f"'{sheet_dep}'!${col_annee_letter}:${col_annee_letter},2023),0)"
            )
            ws.cell(r, 3).number_format = "#,##0"
            ws.cell(r, 3).fill = couleur
            ws.cell(r, 3).border = border_data
            ws.cell(r, 3).alignment = Alignment(horizontal="right")

            ws.cell(r, 4).value = f"=C{r}-B{r}"
            ws.cell(r, 4).number_format = "#,##0"
            ws.cell(r, 4).fill = couleur
            ws.cell(r, 4).border = border_data
            ws.cell(r, 4).alignment = Alignment(horizontal="right")

            ws.cell(r, 5).value = f"=IFERROR((C{r}-B{r})/B{r},0)"
            ws.cell(r, 5).number_format = "0.00%"
            ws.cell(r, 5).fill = couleur
            ws.cell(r, 5).border = border_data
            ws.cell(r, 5).alignment = Alignment(horizontal="right")

            ws.row_dimensions[r].height = 18

        PIE_COL = 27
        PIE_COL_LETTER = get_column_letter(PIE_COL)

        ws.cell(6, PIE_COL).value = "Poste"
        ws.cell(6, PIE_COL + 1).value = "Montant"

        for idx, poste in enumerate(postes_list, start=0):
            r = 7 + idx
            ws.cell(r, PIE_COL).value = poste
            ws.cell(r, PIE_COL + 1).value = (
                f"=IFERROR(SUMIFS('{sheet_dep}'!${col_montant_letter}:${col_montant_letter},"
                f"'{sheet_dep}'!${col_annee_letter}:${col_annee_letter},$B$3,"
                f"'{sheet_dep}'!${col_poste_letter}:${col_poste_letter},$"
                + PIE_COL_LETTER
                + f"${r}),0)"
            )

        chart_pie = PieChart()
        chart_pie.title = "Répartition des Dépenses par Poste"
        chart_pie.height = 15
        chart_pie.width = 20
        chart_pie.style = 12

        data_pie = Reference(
            ws, min_col=PIE_COL + 1, min_row=6, max_row=7 + len(postes_list) - 1
        )
        cats_pie = Reference(
            ws, min_col=PIE_COL, min_row=7, max_row=7 + len(postes_list) - 1
        )

        chart_pie.add_data(data_pie, titles_from_data=True)
        chart_pie.set_categories(cats_pie)

        ws.add_chart(chart_pie, "G10")

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions[PIE_COL_LETTER].hidden = True
        ws.column_dimensions[get_column_letter(PIE_COL + 1)].hidden = True
