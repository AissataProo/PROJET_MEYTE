"""Onglet Sexe/Pathologie : répartition hommes/femmes par pathologie avec graphique empilé 100%."""

from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from config import COULEURS, SHEET_CLEANED_EFFECTIFS


class OngletSexePatho:
    """Affiche la répartition Hommes/Femmes par Pathologie en pourcentage."""

    def __init__(self, wb, df):
        self.wb = wb
        self.df = df
        self.sheet_cleaned = SHEET_CLEANED_EFFECTIFS

    def create(self):
        # 1. Création de l'onglet
        print(f"Colonnes disponibles : {self.df.columns.tolist()}")
        ws = self.wb.create_sheet("Sexe_Pathologie")
        ws.sheet_view.showGridLines = False

        # 2. Titre de l'onglet
        ws["A1"] = "Répartition Hommes/Femmes par Pathologie"
        ws["A1"].font = Font(bold=True, size=14, color=COULEURS["blanc"])
        ws["A1"].fill = PatternFill(
            start_color=COULEURS["principal"], fill_type="solid"
        )
        ws.merge_cells("A1:D1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # 3. Préparation des données avec Pandas (Pivot Table)
        # Utiliser 'patho_niv1' au lieu de 'Pathologie'
        try:
            pivot = self.df.pivot_table(
                values="Effectif",
                index="patho_niv1",  # ✓ Nom correct de colonne
                columns="Sexe",
                aggfunc="sum",
            ).fillna(0)
        except KeyError as e:
            print(f"❌ Erreur: colonne manquante {e}")
            print(f"Colonnes disponibles: {self.df.columns.tolist()}")
            raise

        # 4. Écriture du tableau dans Excel (à partir de la ligne 3)
        rows = dataframe_to_rows(pivot, index=True, header=True)
        for r_idx, row in enumerate(rows, 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # Style pour l'en-tête du tableau
                if r_idx == 3:  # Ligne d'en-tête
                    cell.font = Font(bold=True, color=COULEURS["blanc"], size=10)
                    cell.fill = PatternFill(
                        start_color=COULEURS["principal"], fill_type="solid"
                    )
                elif c_idx == 1:  # Colonne pathologie
                    cell.font = Font(bold=True)
                # Format nombre pour les colonnes Sexe
                if r_idx > 3 and c_idx > 1:
                    cell.number_format = "#,##0"

        # 5. Création du graphique en barres empilées à 100%
        # C'est le meilleur format pour voir la RÉPARTITION (pourcentage)
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "percentStacked"  # Affiche les proportions de 0% à 100%
        chart.overlap = 100
        chart.title = "Répartition Sexe par Pathologie (%)"
        chart.height = 16
        chart.width = 26
        chart.y_axis.title = "Pourcentage (%)"
        chart.x_axis.title = "Pathologie"
        chart.style = 10

        # Sécuriser max_row pour éviter dépassement Excel
        max_row = min(3 + len(pivot) + 1, 1048576)

        data = Reference(
            ws, min_col=2, min_row=3, max_col=pivot.shape[1] + 1, max_row=max_row
        )
        cats = Reference(ws, min_col=1, min_row=4, max_row=max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Couleurs des barres
        chart.series[0].graphicalProperties.solidFill = COULEURS["principal"][2:]

        # Positionnement du graphique à côté du tableau
        ws.add_chart(chart, "E3")

        # 6. Ajustement des colonnes
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15


class OngletAgePatho:
    """Affiche la répartition par Classe d'âge par Pathologie en pourcentage."""

    def __init__(self, wb, df):
        self.wb = wb
        self.df = df
        self.sheet_cleaned = SHEET_CLEANED_EFFECTIFS

    def create(self):
        ws = self.wb.create_sheet("Age_Pathologie")
        ws.sheet_view.showGridLines = False

        # Titre
        ws["A1"] = "Répartition par Classe d'Âge selon Pathologie"
        ws["A1"].font = Font(bold=True, size=14, color=COULEURS["blanc"])
        ws["A1"].fill = PatternFill(
            start_color=COULEURS["principal"], fill_type="solid"
        )
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Pivot table
        pivot = self.df.pivot_table(
            values="Effectif",
            index="patho_niv1",
            columns="Classe d'age",
            aggfunc="sum",
        ).fillna(0)

        # Écriture données
        rows = dataframe_to_rows(pivot, index=True, header=True)
        for r_idx, row in enumerate(rows, 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = Font(bold=True, color=COULEURS["blanc"], size=9)
                    cell.fill = PatternFill(
                        start_color=COULEURS["principal"], fill_type="solid"
                    )
                elif c_idx == 1:
                    cell.font = Font(bold=True)
                if r_idx > 3 and c_idx > 1:
                    cell.number_format = "#,##0"

        # Graphique stacked 100%
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "percentStacked"
        chart.overlap = 100
        chart.title = "Répartition Classe d'Âge par Pathologie (%)"
        chart.height = 16
        chart.width = 28
        chart.y_axis.title = "Pourcentage (%)"
        chart.x_axis.title = "Pathologie"
        chart.style = 11

        max_row = min(3 + len(pivot) + 1, 1048576)
        data = Reference(
            ws,
            min_col=2,
            min_row=3,
            max_col=pivot.shape[1] + 1,
            max_row=max_row,
        )
        cats = Reference(ws, min_col=1, min_row=4, max_row=max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E3")

        ws.column_dimensions["A"].width = 40
        for i in range(2, pivot.shape[1] + 2):
            ws.column_dimensions[chr(64 + i)].width = 12


class OngletDeptPatho:
    """Affiche la répartition par Département selon Pathologie (top 10)."""

    def __init__(self, wb, df):
        self.wb = wb
        self.df = df
        self.sheet_cleaned = SHEET_CLEANED_EFFECTIFS

    def create(self):
        ws = self.wb.create_sheet("Dept_Pathologie")
        ws.sheet_view.showGridLines = False

        # Titre
        ws["A1"] = "Effectifs par Département et Pathologie (Top 15 Depts)"
        ws["A1"].font = Font(bold=True, size=12, color=COULEURS["blanc"])
        ws["A1"].fill = PatternFill(
            start_color=COULEURS["principal"], fill_type="solid"
        )
        ws.merge_cells("A1:G1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Top 15 département par effectif
        top_depts = (
            self.df.groupby("Département")["Effectif"].sum().nlargest(15).index.tolist()
        )

        df_filtered = self.df[self.df["Département"].isin(top_depts)]

        pivot = df_filtered.pivot_table(
            values="Effectif",
            index="Département",
            columns="patho_niv1",
            aggfunc="sum",
        ).fillna(0)

        # Écriture données
        rows = dataframe_to_rows(pivot, index=True, header=True)
        for r_idx, row in enumerate(rows, 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = Font(bold=True, color=COULEURS["blanc"], size=8)
                    cell.fill = PatternFill(
                        start_color=COULEURS["principal"], fill_type="solid"
                    )
                elif c_idx == 1:
                    cell.font = Font(bold=True, size=9)
                if r_idx > 3 and c_idx > 1:
                    cell.number_format = "#,##0"

        # Graphique en barres
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "percentStacked"
        chart.overlap = 100
        chart.title = "Répartition Pathologies par Département (Top 15) (%)"
        chart.height = 18
        chart.width = 28
        chart.y_axis.title = "Pourcentage (%)"
        chart.x_axis.title = "Département"
        chart.style = 12

        max_row = min(3 + len(pivot) + 1, 1048576)
        data = Reference(
            ws,
            min_col=2,
            min_row=3,
            max_col=pivot.shape[1] + 1,
            max_row=max_row,
        )
        cats = Reference(ws, min_col=1, min_row=4, max_row=max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E3")

        ws.column_dimensions["A"].width = 30
        for i in range(2, min(pivot.shape[1] + 2, 10)):
            ws.column_dimensions[chr(64 + i)].width = 12
