"""Onglet Effectifs : tableau de bord effectifs par pathologie et région."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import COULEURS, SHEET_CLEANED_EFFECTIFS
from components.filterseff import add_filter


class OngletEffectifs:
    """Construit l’onglet Excel des effectifs à partir des données nettoyées :
    prépare la feuille de synthèse, applique les filtres utiles et génère les
    tableaux ou visuels permettant d’analyser la répartition des effectifs."""

    def __init__(self, wb, df):
        self.wb = wb
        self.df = df
        self.sheet_cleaned = SHEET_CLEANED_EFFECTIFS

    def create(self):
        ws = self.wb.create_sheet("Effectifs")
        ws.sheet_view.showGridLines = False

        ws_eff = self.wb[self.sheet_cleaned]
        sheet_eff = self.sheet_cleaned

        headers = {}
        for idx, cell in enumerate(ws_eff[1], start=1):
            if cell.value is not None:
                headers[str(cell.value).strip()] = idx

        col_annee = headers.get("annee", 1)
        col_patho = headers.get("patho_niv1", 2)
        col_region = headers.get("Région", 3)
        col_effectif = headers.get("Effectif", 7)

        col_annee_letter = get_column_letter(col_annee)
        col_patho_letter = get_column_letter(col_patho)
        col_region_letter = get_column_letter(col_region)
        col_effectif_letter = get_column_letter(col_effectif)

        annees_list = sorted(self.df["annee"].dropna().unique(), reverse=True)
        if not annees_list:
            annees_list = [2023]

        pathos_list = sorted(self.df["patho_niv1"].dropna().unique())
        if not pathos_list:
            pathos_list = ["Pathologie"]

        regions_list = sorted(self.df["Région"].dropna().unique())
        if not regions_list:
            regions_list = ["Région"]

        patho_defaut = pathos_list[0]
        annee_defaut = int(annees_list[0])
        region_defaut = regions_list[0]

        COLOR_VERT = COULEURS["accent"]
        COLOR_BLEU = COULEURS["secondaire"]

        ws.merge_cells("A1:H1")
        ws["A1"] = "Effectifs par pathologie et région"
        ws["A1"].font = Font(bold=True, size=13, color=COULEURS["blanc"])
        ws["A1"].fill = PatternFill(start_color=COLOR_VERT, fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        add_filter(
            ws, "B2", "Pathologie", pathos_list, patho_defaut, COLOR_BLEU, COLOR_VERT
        )
        add_filter(
            ws,
            "D2",
            "Annee",
            [str(int(a)) for a in annees_list],
            str(int(annee_defaut)),
            COLOR_BLEU,
            COLOR_VERT,
        )
        add_filter(
            ws, "F2", "Région", regions_list, region_defaut, COLOR_BLEU, COLOR_VERT
        )

        TABLE_HDR = 5
        TABLE_START = TABLE_HDR + 1

        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        border_header = Border(
            left=Side(style="thin", color="4472C4"),
            right=Side(style="thin", color="4472C4"),
            top=Side(style="thin", color="4472C4"),
            bottom=Side(style="thin", color="4472C4"),
        )

        couleur_alt_1 = PatternFill(
            start_color=COULEURS["gris_clair"], fill_type="solid"
        )
        couleur_alt_2 = PatternFill(start_color=COULEURS["blanc"], fill_type="solid")

        def make_header_cell(cell, label):
            cell.value = label
            cell.font = Font(bold=True, color=COULEURS["blanc"], size=11)
            cell.fill = PatternFill(start_color=COLOR_BLEU, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_header

        def style_data_cell(cell, idx=0, num_format=None):
            cell.fill = couleur_alt_1 if idx % 2 == 0 else couleur_alt_2
            cell.font = Font(color=COULEURS["noir"])
            cell.border = border
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if num_format:
                cell.number_format = num_format

        make_header_cell(ws.cell(TABLE_HDR, 1), "Région")
        make_header_cell(ws.cell(TABLE_HDR, 2), "Effectifs")
        ws.row_dimensions[TABLE_HDR].height = 22

        for idx, region in enumerate(regions_list):
            r = TABLE_START + idx

            c = ws.cell(r, 1)
            c.value = region
            c.fill = couleur_alt_1 if idx % 2 == 0 else couleur_alt_2
            c.font = Font(color=COULEURS["noir"])
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center")

            c = ws.cell(r, 2)
            c.value = (
                f"=IFERROR(SUMIFS('{sheet_eff}'!${col_effectif_letter}:${col_effectif_letter},"
                f"'{sheet_eff}'!${col_patho_letter}:${col_patho_letter},B$3,"
                f"'{sheet_eff}'!${col_region_letter}:${col_region_letter},A{r},"
                f"'{sheet_eff}'!${col_annee_letter}:${col_annee_letter},D$3),0)"
            )
            style_data_cell(c, idx, "#,##0")

            ws.row_dimensions[r].height = 18

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["F"].width = 20
