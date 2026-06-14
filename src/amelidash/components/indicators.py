"""Filtres interactifs : listes déroulantes avec validation de données Excel."""

from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from config import DEFAULT_PLATFORM2, GREY


def add_filter(
    sheet: Worksheet,
    title: str,
    value: str | int,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
    formula: str,
) -> None:
    """Ajoute un filtre avec liste déroulante à la feuille.

    Le titre et la valeur par défaut sont placés dans des cellules fusionnées
    adjacentes. Pour le filtre "Plateforme", une seconde cellule est ajoutée
    deux lignes en dessous pour permettre la comparaison entre deux plateformes.

    Args:
        sheet: Feuille Excel cible.
        title: Libellé du filtre (ex : "Année", "Plateforme", "Genre").
        value: Valeur sélectionnée par défaut.
        start_row: Première ligne de la plage du titre.
        start_column: Première colonne de la plage du titre.
        end_row: Dernière ligne de la plage du titre.
        end_column: Dernière colonne de la plage du titre. La valeur prend
            les deux colonnes suivantes.
        formula: Référence Excel pour la liste de validation
            (ex : "='Ressources'!$B$1:$B$40").
    """
    _add_filter_title(sheet, title, start_row, start_column, end_row, end_column)
    _add_filter_value(sheet, value, start_row, end_column + 1, end_row, end_column + 2)
    _add_data_validation(sheet, start_row, end_column + 1, formula)

    if title.lower() == "plateforme":
        _add_filter_value(
            sheet,
            DEFAULT_PLATFORM2,
            start_row + 2,
            end_column + 1,
            end_row + 2,
            end_column + 2,
        )
        _add_data_validation(sheet, start_row + 2, end_column + 1, formula)


def _add_filter_title(
    sheet: Worksheet,
    title: str,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> None:
    """Insère et met en forme le libellé d'un filtre avec fusion de cellules.

    Args:
        sheet: Feuille Excel cible.
        title: Texte du libellé.
        start_row: Première ligne de la plage à fusionner.
        start_column: Première colonne de la plage à fusionner.
        end_row: Dernière ligne de la plage à fusionner.
        end_column: Dernière colonne de la plage à fusionner.
    """
    cell = sheet.cell(row=start_row, column=start_column)
    cell.value = title
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=GREY, end_color=GREY, fill_type="solid")
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column,
    )


def _add_filter_value(
    sheet: Worksheet,
    value: str | int,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> None:
    """Insère et met en forme la valeur par défaut d'un filtre avec fusion de cellules.

    Args:
        sheet: Feuille Excel cible.
        value: Valeur par défaut à afficher.
        start_row: Première ligne de la plage à fusionner.
        start_column: Première colonne de la plage à fusionner.
        end_row: Dernière ligne de la plage à fusionner.
        end_column: Dernière colonne de la plage à fusionner.
    """
    cell = sheet.cell(row=start_row, column=start_column)
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=GREY, end_color=GREY, fill_type="solid")
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column,
    )


def _add_data_validation(
    sheet: Worksheet,
    start_row: int,
    start_column: int,
    formula: str,
) -> None:
    """ """
    dv = DataValidation(type="list", formula1=formula)
    sheet.add_data_validation(dv)
    dv.add(sheet.cell(row=start_row, column=start_column).coordinate)
