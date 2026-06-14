from openpyxl.styles import Font, PatternFill, Border, Side

# 1. Configuration centralisée (Idéalement dans config.py)
COULEURS = {
    "principal": "FF004A99",
    "secondaire": "FF008080",
    "accent": "FF4CAF50",
    "filtre": "FFDDEFD8",
    "entete": "FFEFF2F7",
    "texte": "FF333333",
    "alerte": "FFF44336",
    "filtre_bg": "E0E0E0",
}

# 2. Définition des styles (Réutilisables partout)
thin = Side(border_style="thin", color="4D4D4D")
border_thin_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_bottom_only = Border(bottom=thin)
header_fill = PatternFill(start_color="FF4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")


def apply_border_style(sheet, start_row, end_row, start_col, end_col, border_row):
    """
    Applique les bordures sur une plage donnée.
    Si la ligne correspond à border_row, on n'applique qu'une bordure basse.
    """
    for row in sheet.iter_rows(
        min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
    ):
        for cell in row:
            if cell.row == border_row:
                cell.border = border_bottom_only
            else:
                cell.border = border_thin_all
