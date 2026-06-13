"""Filtres interactifs : listes déroulantes avec validation de données Excel."""

from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from config import COULEURS


def add_filter(
    sheet: Worksheet,
    cell: str,
    title: str,
    values: list[str | int],
    default_value: str | int,
    color_value: str,
    color_label: str,
) -> None:
    """Ajoute un filtre avec liste déroulante à la feuille.

    Le titre et la valeur par défaut sont placés dans des cellules fusionnées
    adjacentes.

    Args:
        sheet: Feuille Excel cible.
        cell: Cellule où commencer le filtre (ex : "A1", "B2").
        title: Libellé du filtre (ex : "Année", "Poste", "Pathologie").
        values: Liste des valeurs disponibles dans la liste déroulante.
        default_value: Valeur sélectionnée par défaut.
        color_value: Couleur de fond pour la cellule de valeur.
        color_label: Couleur de fond pour la cellule de titre.
    """
    # Extraire la ligne et la colonne de la cellule (ex: "A1" -> row=1, col=1)
    col_label = cell[0]
    row_num = int(cell[1:])
    col_num = (
        sheet.column_dimensions[col_label].index
        if col_label in sheet.column_dimensions
        else ord(col_label) - ord("A") + 1
    )
    # Titre
    cell_title = sheet.cell(row=row_num, column=col_num)
    cell_title.value = title
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    cell_title.fill = PatternFill(
        start_color=color_label, end_color=color_label, fill_type="solid"
    )
    sheet.merge_cells(
        start_row=row_num, start_column=col_num, end_row=row_num, end_column=col_num
    )
    # Valeur
    col_value_num = col_num + 1
    cell_value = sheet.cell(row=row_num, column=col_value_num)
    cell_value.value = default_value
    cell_value.alignment = Alignment(horizontal="center", vertical="center")
    cell_value.fill = PatternFill(
        start_color=color_value, end_color=color_value, fill_type="solid"
    )
    sheet.merge_cells(
        start_row=row_num,
        start_column=col_value_num,
        end_row=row_num,
        end_column=col_value_num,
    )

    # Validation de données
    formula1 = f'="{",".join(str(v) for v in values)}"'
    dv = DataValidation(type="list", formula1=formula1)
    sheet.add_data_validation(dv)
    dv.add(cell_value)


def _add_filter_title(
    sheet: Worksheet,
    title: str,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> None:
    """Insère et met en forme le libellé d'un filtre avec fusion de cellules.

    Args:
        sheet: Feuille Excel cible.
        title: Texte du libellé.
        start_row: Première ligne de la plage à fusionner.
        start_column: Première colonne de la plage à fusionner.
        end_row: Dernière ligne de la plage à fusionner.
        end_column: Dernière colonne de la plage à fusionner.
    """
    cell = sheet.cell(row=start_row, column=start_column)
    cell.value = title
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(
        start_color=COULEURS["gris_clair"],
        end_color=COULEURS["gris_clair"],
        fill_type="solid",
    )
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column,
    )


def _add_filter_value(
    sheet: Worksheet,
    value: str | int,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> None:
    """Insère et met en forme la valeur par défaut d'un filtre avec fusion de cellules.

    Args:
        sheet: Feuille Excel cible.
        value: Valeur par défaut à afficher.
        start_row: Première ligne de la plage à fusionner.
        start_column: Première colonne de la plage à fusionner.
        end_row: Dernière ligne de la plage à fusionner.
        end_column: Dernière colonne de la plage à fusionner.
    """
    cell = sheet.cell(row=start_row, column=start_column)
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(
        start_color=COULEURS["gris_clair"],
        end_color=COULEURS["gris_clair"],
        fill_type="solid",
    )
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column,
    )


def _add_data_validation(
    sheet: Worksheet,
    start_row: int,
    start_column: int,
    formula: str,
) -> None:
    """Ajoute une validation de données (liste déroulante) à une cellule.

    Args:
        sheet: Feuille Excel cible.
        start_row: Ligne de la cellule à valider.
        start_column: Colonne de la cellule à valider.
        formula: Référence Excel pour la liste de validation
            (ex : "='Ressources'!$B$1:$B$40").
    """
    dv = DataValidation(type="list", formula1=formula)
    sheet.add_data_validation(dv)
    dv.add(sheet.cell(row=start_row, column=start_column).coordinate)